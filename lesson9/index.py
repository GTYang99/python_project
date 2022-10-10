from distutils.cmd import Command
import tkinter as tk
import requests
from tkinter import COMMAND, messagebox
from tkinter import filedialog
import openpyxl

class Window(tk.Tk):
    def __init__(self,codes):
        super().__init__()
        self.codes = codes
        self.title("各縣市4天天氣預測")             # 修改為4天
        print(self.codes)
        tk.Label(self, text="各縣市4天天氣預測", font=("arial",20)).pack(padx=80, pady=30)

        buttonFrame = tk.Frame(self)        
        for index,cities in enumerate(self.codes.items()):            
            cname,ename= cities
            btn = tk.Button(buttonFrame,text=f"{cname}\n{ename}",padx=25,pady=5,width=5)
            btn.grid(column=index % 10,row=index // 10,padx=10,pady=5)
            btn.bind('<Button>',self.btnClick)
        buttonFrame.pack()

        
        self.chage_content_frame = tk.Frame(self) #會改變內容的frame       
        #這裏的內容必需使用者按下按鈕後才面產生
        self.chage_content_frame.pack()

    def btnClick(self,event):
        btn=event.widget
        btn_text = btn["text"]
        nameList = btn_text.split()
        cname = nameList[0]
        ename = nameList[1]
        if hasattr(self, 'displayFrame'):
            self.displayFrame.destroy()
        self.displayFrame = DisplayFrame(self.chage_content_frame,cname=cname,ename=ename)
        self.displayFrame.pack(padx=10,pady=10)

class DisplayFrame(tk.LabelFrame):
    def __init__(self,master,cname,ename):
        super().__init__(master)
        self.cname = cname
        self.ename = ename
        # 刪除列印文字(內部的實體，cname與ename)
        # print(self.cname)
        # print(self.ename)        
        # 新建一個內部實體(self.forecast)，獲取網路上的forecast資料
        # 實體建立一個參數ename
        # self.forecast = DisplayFrame.get_forecast(ename=self.ename)             #下方建立一個檢查，當縣市內沒有資料時，列印發生錯誤，該句關閉
        # 確認是否有獲取到，列印檢查
        # print(self.forecast)
        # print(len(self.forecast))
        #下方建立一個檢查，當縣市內沒有資料時，列印發生錯誤
        try:
            self.forecast = DisplayFrame.get_forecast(ename=self.ename)
        except:
            print("發生錯誤")
            # 利用tkininter中的massagebox建立資料錯誤的訊息窗
            # 但引用這個方法時，為何又需要重新載入tkinter(?)
            messagebox.showerror('取得資料錯誤', '取得資料錯誤,請稍後再試')
            return
        #將資料切割為3等分,left_data,center_data,right_data
        # 1.總資料長度為多少
        total_rows = len(self.forecast)
        # 2.定義行的數量為3
        columns = 3
        # 3.列的數量按照軸跟著軸動態變化，每行3筆資料，數據為40，所以需要多1行
        rows = total_rows // 3 + 1          #40/3=13~1，總計14列
        # 4.1左邊顯示的資料"~到行"
        left_data = self.forecast[:rows]
        # 4.2中間顯示的資料"行到2倍行"
        center_data = self.forecast[rows:rows*2]
        # 4.3右邊顯示的資料"2倍行~"
        right_data = self.forecast[rows*2:]
        # 檢查資料切割數量是否正確
        # print(len(left_data))
        # print(len(center_data))
        # print(len(right_data))
        #建立3欄的顯示空間
        #建立title Label，這段(.configure)不管放在哪都可以出現在顯示區域左上方，是tkinter特定用法
        self.configure(text=f"{self.cname}-{self.ename}")
        # tk.Button(self,text=self.cname).pack()         #分割區域(顏色)時關閉
        # 在tk.button下方顯示一個Frame，按照區域上不同顏色，Frame結尾需要有結束語句
        #左邊顯示空間建立 
        leftFrame = tk.Frame(self,width=150,height=150,borderwidth=1,relief=tk.RIDGE)
        # 新增一個儲存excel功能，需要啟動一個lambda功能，當觸發時會呼叫出DisplayFrame.save_to_excel(self.forecast)內容，self.forecast內容在下方建立一個靜態方法(staticmethod)啟動存檔功能
        # Lambda是匿名函式，是在電腦的編碼過程中，不需要定義函式名稱的一種函式。目前這種函式已經普遍的存在多種程式語言中。
        # 它的功能就像是使用def的方式一樣，只是寫法上更為簡潔，可以寫在行內，在使用前不需要先把函式綁定在一個特定的名稱上。
        tk.Button(self,text=f"{self.cname}天氣預報儲存為Excel",command=lambda:DisplayFrame.save_to_excel(self.forecast,cname)).pack(anchor=tk.W,padx=10,pady=10)
        # 增加行標題
        tk.Label(leftFrame,text="日期-時間",background='#cccccc').grid(row=0,column=0,sticky=tk.W,padx=10,pady=3)
        tk.Label(leftFrame,text="溫度",background='#cccccc').grid(row=0,column=1,sticky=tk.W,padx=10,pady=3)
        tk.Label(leftFrame,text="狀態",background='#cccccc').grid(row=0,column=2,sticky=tk.W,padx=10,pady=3)
        tk.Label(leftFrame,text="濕度",background='#cccccc').grid(row=0,column=3,sticky=tk.W,padx=10,pady=3)
        # 對分割的資料進行遍歷，外迴圈為列，內迴圈為行，機器先確定出行內容，再確定列內容
        for row_index,item in enumerate(left_data):
            for column_index,value in enumerate(item):
                # 原內迴圈為資料文字引用自item中的value，顯示的資料結構按照迴圈次數()來
                # tk.Label(leftFrame,text=value).grid(row=row_index,column=column_index)
                # 修改後內迴圈放的內容需要加上行標題，因此列+1；行不動，資料類型對齊西
                tk.Label(leftFrame,text=value).grid(row=row_index+1,column=column_index,sticky=tk.W,padx=10,pady=3)
        leftFrame.pack(side=tk.LEFT,padx=10)
        #中間顯示空間建立
        centerFrame = tk.Frame(self,width=150,height=150,borderwidth=1,relief=tk.RIDGE)
        # 中間、右邊資料內容按照左邊來修改
        tk.Label(centerFrame,text="日期-時間",background='#cccccc').grid(row=0,column=0,sticky=tk.W,padx=10,pady=3)
        tk.Label(centerFrame,text="溫度",background='#cccccc').grid(row=0,column=1,sticky=tk.W,padx=10,pady=3)
        tk.Label(centerFrame,text="狀態",background='#cccccc').grid(row=0,column=2,sticky=tk.W,padx=10,pady=3)
        tk.Label(centerFrame,text="濕度",background='#cccccc').grid(row=0,column=3,sticky=tk.W,padx=10,pady=3)
        for row_index,item in enumerate(center_data):
            for column_index,value in enumerate(item):
                tk.Label(centerFrame,text=value).grid(row=row_index+1,column=column_index,sticky=tk.W,padx=10,pady=3)
        centerFrame.pack(side=tk.LEFT,padx=10)
        #右邊顯示空間建立
        rightFrame = tk.Frame(self,width=150,height=150,borderwidth=1,relief=tk.RIDGE)
        tk.Label(rightFrame,text="日期-時間",background='#cccccc').grid(row=0,column=0,sticky=tk.W,padx=10,pady=3)
        tk.Label(rightFrame,text="溫度",background='#cccccc').grid(row=0,column=1,sticky=tk.W,padx=10,pady=3)
        tk.Label(rightFrame,text="狀態",background='#cccccc').grid(row=0,column=2,sticky=tk.W,padx=10,pady=3)
        tk.Label(rightFrame,text="濕度",background='#cccccc').grid(row=0,column=3,sticky=tk.W,padx=10,pady=3)
        for row_index,item in enumerate(right_data):
            for column_index,value in enumerate(item):
                tk.Label(rightFrame,text=value).grid(row=row_index+1,column=column_index,sticky=tk.W,padx=10,pady=3)
        # pack指令要寫在實體內容內，上面只有宣告，不能直接使用pack
        rightFrame.pack(side=tk.LEFT,anchor=tk.N,padx=10)
    
    # 建立一個class的method，學習點
    # 基本上class 裡面的function 的第一個參數都需要指向自己(self)，但有兩個例外: staticmethod 與 classmethod
    # 如同大家所熟知的static method，是一種靜態函數。因為是被綁定在類別中，所以不需要建立物件即可使用。也因此他無法存取物件資料，只能對傳入的參數做處理。
    # 什麼時候用static method 比較好？
    # 1.整合工具箱的時候。例如寫一個class專門處理日期格式的轉換，就很適合用static method來處理資料的轉換。
    # 2.只有單一實現的時候。例如你不希望子類別覆寫該函數，為了確保該函數不被覆寫就可加入@staticmethod 這個decorator
    @staticmethod
    # def get_forecast():
    #     return {'key':"value"}
    # 確定上面兩句的方法建立成功，有回傳值後，修改為(lesson8_4內容):
    def get_forecast(ename):        
        url = "https://api.openweathermap.org/data/2.5/forecast?q="+ename+",tw&APPID=46d5b29cf153c79306bb9585d67642ec&lang=zh_tw&units=metric"
        response = requests.request("GET",url)
        if response.ok == True:
            # print("下載成功")    
            all_data = response.json()
        list_data = all_data['list']

        county_forcase = []
        for item in list_data:
            county_forcase.append([item['dt_txt'],item['main']['temp'],item['weather'][0]['description'],item['main']['humidity']])
        return county_forcase
    
    # 建立一個儲存資料形成Excel的靜態方法
    @staticmethod
    def save_to_excel(data,cname):    
        # 檢查資料是否可以正確導入到靜態方法中
        # print("儲存資料至excel")
        # print(data)
        # 使用openpxyl對資料儲存
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "天氣預測表"
        sheet.cell(row=1, column=1, value="日期-時間")
        sheet.cell(row=1, column=2, value="溫度")
        sheet.cell(row=1, column=3, value="狀態")
        sheet.cell(row=1, column=4, value="濕度")
        # 巢狀迴圈，for i , value in enumerate(list)，取得list的在第i位是value元素(標準用法，學習點)
        for row_index,row_data in enumerate(data):
            for column_index,value in enumerate(row_data):
                sheet.cell(row=row_index+2,column=column_index+1,value=value)
        # 自動選擇為桌面存檔
        filePath = filedialog.askdirectory()
        # 增加一個自動出現檔案名稱功能，引用自父類別的ename
        # print(ename)
        wb.save(filePath+f'/氣象預測_{cname}.xlsx')
        messagebox.showinfo("ok",message="存檔成功")
        

def main():
    tw_county_names = {"台北":"Taipei",
                   "台中":"Taichung",
                   "基隆":"Keelung",
                   "台南":"Tainan",
                   "高雄":"Kaohsiung",
                   "新北":"New Taipei",
                   "宜蘭":"Yilan",
                   "桃園":"Taoyuan",
                   "嘉義":"Chiayi",
                   "新竹":"Hsinchu",
                   "苗栗":"Miaoli",
                   "南投":"Nantou",
                   "彰化":"Changhua",
                   "雲林":"Yunlin",
                   "屏東":"Pingtung",
                   "花蓮":"Hualien",
                   "台東":"Taitung",
                   "金門":"Kinmen",
                   "澎湖":"Penghu",
                   "連江":"Lienchiang"
                   }
    window = Window(codes=tw_county_names)
    window.mainloop()

if __name__ == "__main__":
    main()